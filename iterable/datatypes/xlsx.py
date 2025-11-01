from __future__ import annotations
import typing
from openpyxl import load_workbook

from ..base import BaseFileIterable, BaseCodec

def read_row_keys(rownum, ncols, sheet):
    """Read single row by row num"""
    tmp = list()
    for i in range(0, ncols):
        ct = sheet.cell_type(rownum, i)
        cell_value = sheet.cell_value(rownum, i)
        get_col = str(cell_value)
        tmp.append(get_col)    
    return tmp


class XLSXIterable(BaseFileIterable):
    datamode = 'binary'
    def __init__(self, filename:str = None, stream:typing.IO = None, codec: BaseCodec = None, mode='r', keys: list[str] = None, page:int = 0, start_line:int = 0, options:dict={}):
        super(XLSXIterable, self).__init__(filename, stream, codec=codec, mode=mode, binary=True, noopen=True, options=options)
        self.keys = keys
        self.start_line = start_line + 1
        self.page = page
        if 'page' in options.keys():
            self.page = int(options['page'])
        self.pos = self.start_line
        self.extracted_keys = keys is None
        self.reset()
        pass

    def reset(self):
        super(XLSXIterable, self).reset()
        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook[self.workbook.sheetnames[self.page]]
        self.pos = self.start_line
        self.cursor = self.sheet.iter_rows()
        if self.pos > 1:
            self.skip(self.pos - 1)
        if self.extracted_keys:
            self.keys = list()
            row = next(self.cursor)
            for cell in row:
                self.keys.append(str(cell.value))             
            self.pos += 1

    def skip(self, num:int = 1):
        while num > 0:
            num -= 1
            o = next(self.cursor)

    @staticmethod
    def has_totals():
        """Has totals indicator"""
        return True        

    def totals(self):
        """Returns file totals"""
        return self.sheet.max_row

    @staticmethod
    def id() -> str:
        return 'xlsx'

    @staticmethod
    def is_flatonly() -> bool:
        return True

    def read(self) -> dict:
        """Read single XLSX record"""
        row = next(self.cursor)
        tmp = list()
        for cell in row:
            tmp.append(str(cell.value))
        result = dict(zip(self.keys, tmp))
        self.pos += 1
        return result

    def read_bulk(self, num:int = 10) -> list[dict]:
        """Read bulk XLSX records"""
        chunk = []
        for n in range(0, num):
            row = next(self.cursor)
            tmp = list()
            for cell in row:
                tmp.append(str(cell.value))
            result = dict(zip(self.keys, tmp))
            chunk.append(result)
            self.pos += 1
        return chunk
